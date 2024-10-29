# Copyright (c) 2015, Frappe Technologies Pvt. Ltd. and Contributors
# License: MIT. See LICENSE
import re
from io import BytesIO

import openpyxl
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell
from openpyxl.workbook.child import INVALID_TITLE_REGEX

import frappe
from frappe.utils.html_utils import unescape_html

ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


# return xlsx file object
def make_xlsx(data, sheet_name, wb=None, column_widths=None, column_formats=None, freeze=None):
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook(write_only=True)

	sheet_name_sanitized = INVALID_TITLE_REGEX.sub(" ", sheet_name)
	ws = wb.create_sheet(sheet_name_sanitized, 0)

	if freeze:
		ws.freeze_panes = freeze

	# Column widths
	for i, column_width in enumerate(column_widths):
		if column_width:
			ws.column_dimensions[get_column_letter(i + 1)].width = column_width

	# Column formats
	column_formats = column_formats or []
	for i, column_format in enumerate(column_formats):
		if column_format:
			ws.column_dimensions[get_column_letter(i + 1)].number_format = column_format

	# Header Row Bold
	row1 = ws.row_dimensions[1]
	row1.font = Font(name="Calibri", bold=True)

	for row_i, row in enumerate(data):
		clean_row = []
		for col_i, item in enumerate(row):
			if isinstance(item, str) and (sheet_name not in ["Data Import Template", "Data Export"]):
				value = handle_html(item)
			else:
				value = item

			if isinstance(item, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
				# Remove illegal characters from the string
				value = ILLEGAL_CHARACTERS_RE.sub("", value)

			cell = WriteOnlyCell(ws, value)
			cell.number_format = ws.column_dimensions[get_column_letter(col_i + 1)].number_format
			if row_i == 0:
				cell.font = Font(name="Calibri", bold=True)

			clean_row.append(cell)

		ws.append(clean_row)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file


def handle_html(data):
	from frappe.core.utils import html2text

	# return if no html tags found
	data = frappe.as_unicode(data)

	if "<" not in data or ">" not in data:
		return data

	h = unescape_html(data or "")

	try:
		value = html2text(h, strip_links=True, wrap=False)
	except Exception:
		# unable to parse html, send it raw
		return data

	value = ", ".join(value.split("  \n"))
	value = " ".join(value.split("\n"))
	return ", ".join(value.split("# "))


def read_xlsx_file_from_attached_file(file_url=None, fcontent=None, filepath=None):
	if file_url:
		_file = frappe.get_doc("File", {"file_url": file_url})
		filename = _file.get_full_path()
	elif fcontent:
		filename = BytesIO(fcontent)
	elif filepath:
		filename = filepath
	else:
		return

	rows = []
	wb1 = load_workbook(filename=filename, data_only=True)
	ws1 = wb1.active
	for row in ws1.iter_rows():
		rows.append([cell.value for cell in row])
	return rows


def read_xls_file_from_attached_file(content):
	book = xlrd.open_workbook(file_contents=content)
	sheets = book.sheets()
	sheet = sheets[0]
	return [sheet.row_values(i) for i in range(sheet.nrows)]


def build_xlsx_response(data, filename):
	from frappe.desk.utils import provide_binary_file

	provide_binary_file(filename, "xlsx", make_xlsx(data, filename).getvalue())
